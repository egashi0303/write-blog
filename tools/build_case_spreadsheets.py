from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE = Path(r"C:\Users\user\write-blog\outputs")
THUMBNAIL_PREP_BASE = BASE / "記事生成の準備"


def load_body(path: Path) -> str:
    body = path.read_text(encoding="utf-8").strip()
    if body.startswith("# "):
        body = "\n".join(body.splitlines()[1:]).lstrip()
    return body


def build_case_workbook(output_path: Path, date: str, case_id: str, case_name: str, angle: str, title: str, body_path: Path) -> None:
    body = load_body(body_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "記事一覧"
    headers = ["記事作成日", "案件ID", "案件名", "切り口", "記事タイトル", "記事本文", "サムネイルURL"]
    sheet.append(headers)
    sheet.append([date, case_id, case_name, angle, title, body, ""])

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {"A": 14, "B": 12, "C": 20, "D": 24, "E": 48, "F": 90, "G": 28}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.row_dimensions[1].height = 24
    workbook.save(output_path)
    prepare_thumbnail_artifacts(body_path, title)


def extract_title(path: Path) -> str:
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith("# "):
            return line[2:].strip()
    return path.stem


def sanitize_name(name: str) -> str:
    invalid = '<>:"/\\|?*'
    for char in invalid:
        name = name.replace(char, "_")
    return name.strip()[:180]


def build_prompt(title: str, article_path: Path) -> str:
    return f"""# Thumbnail Prompt

Source article: `{article_path}`

## Title

{title}

## Prompt

Create a photorealistic editorial photo, not an illustration, not a vector graphic, not a 3D render, not a cartoon.
Use a blank, text-zero canvas with no typography, no caption blocks, no title blocks, and no labels.
Preferred canvas type: desktop_wallpaper.

Main subject: [記事内容に合う具体的な被写体]

Article theme: {title}

Setting: warm natural light, realistic environment, subtle relevant props, calm and trustworthy atmosphere.

Leave clean negative space for later Japanese title overlay, especially upper-left and center-left.

Do not include any readable text, letters, numbers, logos, brand names, or fake UI text.
Avoid flashy colors, luxury imagery, stacks of cash, coins, aggressive screens, anxiety, and gambling feeling.
Professional blog thumbnail photograph, soft contrast, warm but restrained colors.
"""


def prepare_thumbnail_artifacts(article_path: Path, fallback_title: str) -> None:
    title = extract_title(article_path) or fallback_title
    folder = THUMBNAIL_PREP_BASE / sanitize_name(article_path.stem)
    folder.mkdir(parents=True, exist_ok=True)

    (folder / "thumbnail_prompt.md").write_text(build_prompt(title, article_path), encoding="utf-8")
    (folder / "thumbnail_generation_report.json").write_text(
        json.dumps(
            {
                "mvp_scope": "title_to_canva_single_thumbnail_background",
                "template_policy": {
                    "name": "thumbnail_no_text_template",
                    "must_use_every_time": True,
                    "preferred_canva_design_type": "desktop_wallpaper",
                    "reject_if_any_readable_text_appears": True,
                },
                "source_article": str(article_path),
                "source_title": title,
                "canva_design_type": "desktop_wallpaper",
                "canva_job_id": "",
                "generated_candidate_count": 0,
                "selected_policy": "first_candidate_for_mvp",
                "selected_candidate_id": "",
                "selected_candidate_preview_url": "",
                "selected_candidate_thumbnail_url": "",
                "created_design": {
                    "id": "",
                    "title": "",
                    "edit_url": "",
                    "view_url": "",
                    "page_count": 0,
                },
                "download_status": "not_started",
                "local_download_path": "",
                "drive_upload": {
                    "folder_id": "",
                    "folder_name": "",
                    "folder_url": "",
                    "file_id": "",
                    "file_name": "",
                    "file_url": "",
                },
                "constraints": {
                    "photorealistic_only": True,
                    "no_text_template": True,
                    "text_overlay_added": False,
                },
                "notes": [
                    "記事タイトルを読み取ってサムネイル準備を開始するための下書き。",
                    "Canva生成後は thumbnail_url を使ってDrive保存する。",
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    (folder / "next_step.md").write_text(
        "\n".join(
            [
                "# Next Step",
                "",
                f"1. Read title: {title}",
                "2. Use the no-text thumbnail prompt from `thumbnail_prompt.md`",
                "3. Generate Canva candidates with `desktop_wallpaper`",
                "4. Pick a candidate with no readable text",
                "5. Fetch the thumbnail URL from Canva",
                "6. Publish with `tools/publish_thumbnail_report_to_drive.mjs`",
            ]
        ),
        encoding="utf-8",
    )


def main() -> None:
    build_case_workbook(
        BASE / "株案件テスト.xlsx",
        "2026-06-03",
        "CASE-0005",
        "株案件テスト",
        "株式投資の基本を知る",
        "60代からでも遅くない。株式投資の前に知っておきたい基本と考え方",
        BASE / "generated_articles" / "CASE-0005_20260603_stock_basics.md",
    )
    build_case_workbook(
        BASE / "サロンドオズ.xlsx",
        "2026-06-02",
        "CASE-0003",
        "サロンドオズ",
        "脱毛方法の選び方",
        "脱毛方法で迷う方へ。WAX・シュガーリング・ルミクスを知って、自分に合うケアを選ぶ",
        BASE / "generated_articles" / "CASE-0003_20260602_salondoz.md",
    )
    build_case_workbook(
        BASE / "medinaturals.xlsx",
        "2026-06-02",
        "CASE-0004",
        "medinaturals",
        "CBDを回復の選択肢として知る",
        "眠りの浅さや疲れが気になるとき、CBDを選択肢のひとつに",
        BASE / "generated_articles" / "CASE-0004_20260602_cbd_recovery.md",
    )
    print("created case spreadsheets")


if __name__ == "__main__":
    main()
